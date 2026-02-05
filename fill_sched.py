from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re


# цвет, которым будут закрашиваться клетки
changed_fill = PatternFill(
    start_color="FFFF99",
    end_color="FFFF99",
    fill_type="solid"
)

# нормализуем строку для её проверки
def normal(value):
    if value is None:
        return ""
    
    # убираем лишние пакости и в целом приводим к нормальной строке
    text = re.sub(r"[\s\-—_,/]+", " ", str(value))
    text = text.replace(" ", "")
    return text.lower()

# функция закраски клетки
def fill_sched(old_sched, new_sched):
    
    old_wb = load_workbook(old_sched)
    new_wb = load_workbook(new_sched)
    
    old_ws = old_wb.active
    new_ws = new_wb.active
    
    # проверка по совпадению строк
    for row in range(1, new_ws.max_row + 1):
        for col in range(1, new_ws.max_column + 1):
            old_cell = old_ws.cell(row=row, column=col)
            new_cell = new_ws.cell(row=row, column=col)
            
            old_val = normal(old_cell.value)
            new_val = normal(new_cell.value)
                
            if old_val != new_val:
                new_cell.fill = changed_fill
            
    new_wb.save("sched.xlsx")