from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from settings import global_classes

# цвет, которым будут закрашиваться клетки. так же используется для проверки
changed_fill = PatternFill(
    start_color="FFFF99",
    end_color="FFFF99",
    fill_type="solid"
)


# получаем день расписания
def get_day(file='sched.xlsx'):
    
    wb = load_workbook(file)
    ws = wb.active

    return ws.cell(row=1, column=1).value.split("\n")[1]


# функция нахождения номера урока
def find_num_col(ws, class_row, class_col, att=1, dist=False) -> int:

    # идём влево по колонкам и на 1 вниз
    for c in range(class_col - 1, 0, -1):
        v = ws.cell(row=class_row + att, column=c).value

        # устраиваем проверку на то, является ли значение номером урока
        if "дистан" in normal(v):
            return find_num_col(ws, class_row, class_col, att=att+1, dist=True)
        if is_lesson_num(v):
            return c, class_row + att, dist

    print('Не найден номер урока')
    return None


# функция проверки клетки на наличие номера урока
def is_lesson_num(v) -> bool:
    # на всякий проверяем если вместо None пустая клетка будет содержать пробелы
    if normal(v) == '':
        return False
    s = str(v).strip()

    return s.isdigit()


# нормализуем строку для её проверки
def normal(value):
    if value is None:
        return ""
    
    # убираем лишние пакости и в целом приводим к нормальной строке
    text = re.sub(r"[\s\-—_,/<>]+", " ", str(value))
    text = text.replace(" ", "")
    return text.lower()


# нормализуем строку для отправки
def normal_4_send(value):
    if value is None or normal(value) == "":
        return " -"
    
    # на случай тире в названии урока и переносах строки делаем так
    text = value.replace("-", '').replace("—", '').replace("\n", '')
    text = re.sub(r"[\s<>]+", " ", text)
    return text


# функция закраски клетки
def fill_sched(old_sched, new_sched):
    
    old_wb = load_workbook(old_sched)
    new_wb = load_workbook(new_sched)
    
    old_ws = old_wb.active
    new_ws = new_wb.active
    
    # проверка по совпадению строк
    for row in range(1, new_ws.max_row + 1):
        for col in range(1, new_ws.max_column + 1):
            old_cell = old_ws.cell(row=row, column=col)
            new_cell = new_ws.cell(row=row, column=col)
            
            old_val = normal(old_cell.value)
            new_val = normal(new_cell.value)
                
            if old_val != new_val:
                new_cell.fill = changed_fill
            
    new_wb.save("sched.xlsx")


# функция поиска расписания для каждого класса
def find_in_sched(new=True, file='sched.xlsx', classes=global_classes, get=False):
    
    wb = load_workbook(file)
    ws = wb.active

    if classes[0] not in global_classes:
        return f'Класс {classes[0].upper()} не найден!'

    # устраиваем перебор по названию классов в списке
    for class_name in classes:

        class_row = None
        class_col = None

        # перебираем всю таблицу для поиска нужного нам класса
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws_cell = ws.cell(row=row, column=col)

                if normal(ws_cell.value) == class_name:
                    class_row = row
                    class_col = col
                    break
            
            # если нашли класс, то ломаем цикл
            if class_col:
                break

        # если не нашли класс
        if not class_col:
            print(f"Класс {class_name} не найден")
            continue

        # нашли класс, потом переменная будет использоваться для отправки всем пользователям по ней 
        class_name = (f"{class_name.upper()}")

        # список уроков с классом сверху
        day = ws.cell(row=1, column=1).value.split("\n")[1]
        lessons = f'<b>{day}</b><b>\n\n{class_name}</b>\n\n'

        # ищем колонку с номером
        num_col, r, dist = find_num_col(ws, class_row, class_col)

        if dist:
            lessons += "<b>Дистанционное обучение!</b>\n\n"

        if not num_col:
            print("не найдена колонка с номерами")
            continue

        # переменная для изменённого расписания. для нового не сыграет никакой роли
        has_lessons = False

        # перебираем уроки
        while True:
            # номер урока, сама клетка (для проверки по цвету), урок
            lesson_num = ws.cell(row=r, column=num_col).value
            lesson_time = ws.cell(row=r, column=num_col + 1).value
            lesson_cell = ws.cell(row=r, column=class_col)
            lesson = lesson_cell.value

            # если номер урока закончился, то выходим из цикла
            if not is_lesson_num(lesson_num):
                break
            
            lesson = f'{normal_4_send(lesson)}'
            # если расписание не новое, т.е. изменённое, то проходим проверку
            if not new:
                # цвет клетки
                fill = lesson_cell.fill

                if fill == changed_fill:
                    has_lessons = True
                    lessons += f'<blockquote>{lesson_num}. <b>(Изм.)</b> {lesson}</blockquote>\n<i>{lesson_time}</i>\n'
                else:
                    lessons += f'<blockquote>{lesson_num}. {lesson}</blockquote>\n<i>{lesson_time}</i>\n'

            else:
                lessons += f'<blockquote>{lesson_num}. {lesson}</blockquote>\n<i>{lesson_time}</i>\n'

            # идём дальше вниз на ещё одну клетку
            r += 1

        # lessons - уроки для одного класса. выводит его если расписание новое, либо если в уроках класса есть изменения
        if has_lessons or new or get:
            return lessons
        else:
            return None